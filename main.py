import sys
import re
import io
from datetime import datetime
from decimal import Decimal
from ofxparse import OfxParser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def pegar_dados(lnkObjeto, lnkAtributo, lnkPadrao=None):
    return getattr(lnkObjeto, lnkAtributo, lnkPadrao) if lnkObjeto is not None else lnkPadrao

def formatar_data(lnkData):
    if not lnkData:
        return None
    if isinstance(lnkData, (datetime, )):
        return lnkData.strftime("%d-%m-%Y")
    
    try:
        return lnkData.strftime("%d-%m-%Y")
    except Exception:
        return str(lnkData)

def pegar_totais(lnkObjeto):
    transacoes = pegar_dados(lnkObjeto, "transactions", []) or []
    totalCredito = Decimal("0")
    totalDebito = Decimal("0")

    for umaTransacao in transacoes:
        total = Decimal(str(pegar_dados(umaTransacao, "amount", 0) or 0))
        if total >= 0:
            totalCredito += total
        else:
            totalDebito += total
    
    return {
        "total_credito" : totalCredito,
        "total_debito" : totalDebito,
        "quantidade_transacoes" : len(transacoes),
    }

def pegar_dados_conta(lnkConta):
    instituicao = pegar_dados(lnkConta, "institution")
    declaracao = pegar_dados(lnkConta, "statement")
    
    banco = pegar_dados(lnkConta, "bank_id")
    if not banco:
        banco = pegar_dados(instituicao, "fid") or pegar_dados(instituicao, "org")

    agencia = pegar_dados(lnkConta, "branch_id") or pegar_dados(lnkConta, "routing_number")
    numeroConta = pegar_dados(lnkConta, "account_id")
    tipoConta = pegar_dados(lnkConta, "account_type")

    match tipoConta:
        case "CHECKING":
            tipoConta = "Conta corrente"
        case "SAVINGS":
            tipoConta = "Poupança"
        case "CREDITCARD":
            tipoConta = "Cartão de crédito"
    
    dataInicial = formatar_data(pegar_dados(declaracao, "start_date"))
    dataFinal = formatar_data(pegar_dados(declaracao, "end_date"))
    dataSaldo = formatar_data(pegar_dados(declaracao, "balance_date"))
    saldo = pegar_dados(declaracao, "balance")
    saldoDisponivel = pegar_dados(declaracao, "available_balance")

    dados = {
        "banco" : banco,
        "tipo" : tipoConta,
        "agencia" : agencia,        
        "conta" : numeroConta,
        "saldo" : saldo,
        "saldo_disponivel" : saldoDisponivel,
        "data_inicial" : dataInicial,
        "data_final" : dataFinal,
        "nome_banco" : pegar_dados(instituicao, "name"),
        "organizacao" : pegar_dados(instituicao, "org"),
    }

    dados.update(pegar_totais(declaracao))

    return dados


def normalizar_cabecalho(lnkTexto):
    """
    Corrige casos de erro de encoding, como por exemplo:
    Banco C6: UTF - 8
    """
    def _compactar(t):
        # Remove todos os espaços na parte capturada (UTF - 8 -> UTF-8)
        valor = re.sub(r"\s+", "", t.group(2))
        return t.group(1) + valor

    # Flags re.IGNORECASE para pegar variações
    lnkTexto = re.sub(r"(ENCODING:\s*)([^\r\n]+)", _compactar, lnkTexto, flags=re.IGNORECASE)
    lnkTexto = re.sub(r"(CHARSET:\s*)([^\r\n]+)", _compactar, lnkTexto, flags=re.IGNORECASE)
    return lnkTexto.lstrip("\ufeff")  # remove BOM se houver

def processar_arquivo(lnkArquivoOfx, lnkArquivoExcel):

    conteudo = open(lnkArquivoOfx, "rb").read()

    #Pegando só começo do arquivo
    cabecalho = conteudo[:4096].decode("ascii", errors="ignore")
    encodeDeclarado = re.search(r"ENCODING:\s*([^\r\n]+)", cabecalho, flags=re.I)
    charsetDeclarado = re.search(r"CHARSET:\s*([^\r\n]+)", cabecalho, flags=re.I)

     # Ajusta encoding conforme dados e cabeçalho
    encodeDeclarado = re.sub(r"\s+", "", encodeDeclarado.group(1)) if encodeDeclarado else None
    charsetDeclarado = re.sub(r"\s+", "", charsetDeclarado.group(1)) if charsetDeclarado else None

    isTemCaracterMaior = any(b >= 0x80 for b in conteudo)
    encode = "utf-8"
    
    if encodeDeclarado in ("USASCII", "ASCII"):
        encode = "cp1252" if isTemCaracterMaior else "ascii"
    elif encodeDeclarado in ("ISO-8859-1", "LATIN1", "LATIN-1"):
        encode = "latin-1"
    elif encodeDeclarado in ("UTF-8", "UTF8"):
        encode = "utf-8"
    elif charsetDeclarado == "1252":
        encode = "cp1252"

    #Decodificando o conteúdo
    for i in (encode, "utf-8-sig", "cp1252", "latin-1"):
        try:
            textoDecodificado = conteudo.decode(i)
            break
        except UnicodeDecodeError:
            continue
    else:
        textoDecodificado = conteudo.decode("utf-8", errors="ignore")
    
    #Alguns bancos estão retornando o encoding fora do padrão, como o banco C6 por exemplo, que retornou o enconding = UTF - 8    
    conteudoCorrigido = normalizar_cabecalho(textoDecodificado)
    
    # Usa StringIO para forçar "arquivo de texto" seguro no parser
    conteudoOfx = io.StringIO(conteudoCorrigido)
    
    #Parseando a string em um objeto
    ofx = OfxParser.parse(conteudoOfx)

    #Criando uma planilha no formato Microsoft Excel
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Transacoes"

    cabecalho = ["ID","Banco","Conta", "Data", "Tipo", "Valor", "Descrição", "Beneficiario", "Cheque"]
    sheet.append(cabecalho)

    #Estilizando o cabeçalho
    for coluna in range(1, len(cabecalho) + 1 ):
        celula = sheet.cell(row=1, column=coluna)
        celula.font = Font(bold=True)
        celula.alignment = Alignment(horizontal="center")
        celula.number_format="R$ #,##0.00_);[Red](R$ #,##0.00)"
    
    #Lendo as informações de transação
    for conta in ofx.accounts:
        
        dados = pegar_dados_conta(conta) 
        banco = dados["banco"]
        numeroConta = dados["conta"]

        tamanhoMemo = 0
        tamanhoBeneficiario = 0
        tamanhoID = 0
        
        linha = 1
        for transacao in conta.statement.transactions:
            id = transacao.id
            data = formatar_data(transacao.date)
            tipo = transacao.type
            valor = transacao.amount
            memo = transacao.memo
            beneficiario = transacao.payee
            cheque = transacao.checknum

            match tipo:
                case "credit":
                    tipo = "CRÉDITO"
                case "debit":
                    tipo = "DÉBITO"

            tamanhoBeneficiario = len(str(beneficiario)) if len(str(beneficiario)) > tamanhoBeneficiario else tamanhoBeneficiario
            tamanhoMemo = len(str(memo)) if len(str(memo)) > tamanhoMemo else tamanhoMemo
            tamanhoID = len(str(id)) if len(str(id)) > tamanhoID else tamanhoID

            sheet.append([id, banco, numeroConta, data, tipo, valor, memo, beneficiario, cheque])

            linha+= 1
            celula = sheet.cell(row=linha, column=6)
            celula.number_format = 'R$ #,##0.00_);[Red]-R$ #,##0.00'


    
    #Ajustando o tamanho das colunas
    #A biblioteca openpyxl não possui um AutoFit, por isso precisamos deixar a coluna com uma largura fixa
    #Outra solução seria calcular o comprimento do conteúdo mais longo da coluna, e acredito não ser necessário para esse caso.
    tamanhoBeneficiario+= 2
    tamanhoMemo+= 2
    tamanhoID += 2

    tamanhoColunas = [tamanhoID,7,12,12,10,12,tamanhoMemo,tamanhoBeneficiario,20]
    for i, tamanhoDesejado in enumerate(tamanhoColunas, start=1):
        sheet.column_dimensions[chr(64 + i)].width = tamanhoDesejado

    ##Salvando o arquivo processado
    wb.save(lnkArquivoExcel)


#Validando o número máximo de parâmetros necessários para execução da função.
if len(sys.argv) != 3:
    print("Este programa precisa de 02 parâmetros para execução! \n Parâmetro 01: arquivo ofx \n Parâmetro 02: caminho completo para retorno do processamento.")
    sys.exit(1)

#Captando as informações de entrada
arquivo_ofx = sys.argv[1]
arquivo_excel = sys.argv[2]

if not arquivo_ofx.lower().endswith(".ofx"):
    print("Extensão inválida, aguardando um arquivo com extensão OFX. Arquivo informado: ", arquivo_ofx)
    sys.exit(1)

#Chamando a rotina principal
processar_arquivo(arquivo_ofx, arquivo_excel)